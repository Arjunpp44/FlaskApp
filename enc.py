import time
import os
import openpyxl
# from crypto.Cipher import AES
# from crypto.Util.Padding import pad, unpad
from Crypto.Cipher import AES
from crypto.Util.Padding import unpad, pad
from openpyxl import Workbook

def aesEncrypt(data,key,cipherIV):
    try:
            plaintext = data.encode('utf8')
            cipher = AES.new(key.encode('utf8'), AES.MODE_CBC, bytes.fromhex(cipherIV))
            ciphertext = cipher.encrypt(pad(plaintext, AES.block_size))
            return ciphertext.hex()
    except:
        print('error')
        return ''





def aesDecrypt(ciphertext,key,cipherIV):
    try:
        cipher = AES.new(key.encode('utf8'), AES.MODE_CBC, bytes.fromhex(cipherIV))
        plaintext = unpad(cipher.decrypt(bytes.fromhex(ciphertext)), AES.block_size)
        return plaintext.decode('utf8')
    except:
        print('decypt error')



def get_data_from_excel(filepath, sheet_name):
    behaviour_exel_data = openpyxl.load_workbook(filepath)
    sheet_obj = behaviour_exel_data.get_sheet_by_name(sheet_name)
    if not sheet_obj:
        return False, 'Invalid sheet name'
    result = []
    for row_item in range(2, sheet_obj.max_row + 1):
        dict_val = {}
        for col_item in range(1, sheet_obj.max_column + 1):
            cell = sheet_obj.cell(row=row_item, column=col_item).value
            col_val = sheet_obj.cell(row=1, column=col_item).value
            if col_val:
                try:
                    dict_val[col_val] = str(int(cell))
                except Exception as e:
                    dict_val[col_val] = cell

        if set(dict_val.values()) != {None}:
            result.append(dict_val)
    return True, result

def create_quick_exel(data, file_base_name):
    wb = Workbook()
    ws = wb.active
    heading_list = []
    for d in data:
        for k in d.keys():
            if k not in heading_list:
                heading_list.append(k)
    heading_count = 1
    for headings in heading_list:
        ws.cell(row=1, column=heading_count).value = headings
        heading_count += 1

    for i in range(len(data)):
        camp_data = data[i]
        index = i + 2
        column = 1
        for heading in heading_list:
            ws.cell(row=index, column=column).value = camp_data.get(heading, "")
            column += 1
    curr_time = "a"+str(time.time())
    save_path=os.path.join('C:\\Users\\remshad\\PycharmProjects\\flaskPdfKit','inpnjik')
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    file_path = os.path.join(save_path, f'{file_base_name}{curr_time}.xlsx')
    print(file_path)
    wb.save(file_path)
    return file_path


